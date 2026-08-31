"""
common.py - Shared utility functions for AMD static analysis pipeline.
Provides: hash computation, entropy calculation, APK zip operations,
null cleaning, resource loading, and jadx-gui lifecycle management.
"""
import os
import sys
import json
import math
import re
import zipfile
import hashlib
import subprocess
import time
import threading
import signal


# ── Hash & Entropy ────────────────────────────────────────────────────────────

def compute_file_hashes(file_path: str) -> dict:
    """Compute MD5, SHA1, SHA256 for a file. Returns dict with md5, sha1, sha256 keys."""
    h_md5 = hashlib.md5()
    h_sha1 = hashlib.sha1()
    h_sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h_md5.update(chunk)
            h_sha1.update(chunk)
            h_sha256.update(chunk)
    return {
        "md5": h_md5.hexdigest(),
        "sha1": h_sha1.hexdigest(),
        "sha256": h_sha256.hexdigest(),
    }


def shannon_entropy_bytes(data: bytes) -> float:
    """Shannon entropy of raw bytes (0-8 range)."""
    if not data:
        return 0.0
    freq = [0] * 256
    for b in data:
        freq[b] += 1
    length = len(data)
    ent = 0.0
    for count in freq:
        if count > 0:
            p = count / length
            ent -= p * math.log2(p)
    return round(ent, 4)


def shannon_entropy_string(s: str) -> float:
    """Shannon entropy of a character string."""
    if not s:
        return 0.0
    freq = {}
    for ch in s:
        freq[ch] = freq.get(ch, 0) + 1
    length = len(s)
    ent = 0.0
    for count in freq.values():
        if count > 0:
            p = count / length
            ent -= p * math.log2(p)
    return round(ent, 4)


# ── APK operations ────────────────────────────────────────────────────────────

def extract_apk(apk_path: str, output_dir: str) -> str:
    """Extract APK to directory. Returns extraction directory path."""
    os.makedirs(output_dir, exist_ok=True)
    with zipfile.ZipFile(apk_path, "r") as zf:
        zf.extractall(output_dir)
    return output_dir


def list_apk_files(apk_path: str) -> list:
    """List all files inside APK."""
    with zipfile.ZipFile(apk_path, "r") as zf:
        return zf.namelist()


def read_apk_file_bytes(apk_path: str, inner_path: str) -> bytes:
    """Read a specific file from APK as bytes."""
    with zipfile.ZipFile(apk_path, "r") as zf:
        return zf.read(inner_path)


def get_dex_files(apk_path: str) -> list:
    """Get list of DEX files (classes.dex, classes2.dex, ...) in APK."""
    dex_list = []
    with zipfile.ZipFile(apk_path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("classes") and name.endswith(".dex"):
                dex_list.append(name)
    return dex_list


def get_native_libs(apk_path: str) -> list:
    """Get list of .so files in APK."""
    so_list = []
    with zipfile.ZipFile(apk_path, "r") as zf:
        for name in zf.namelist():
            if name.endswith(".so"):
                so_list.append(name)
    return so_list


def compute_file_entropy_in_apk(apk_path: str, inner_path: str) -> float:
    """Compute Shannon entropy for a file inside the APK."""
    try:
        data = read_apk_file_bytes(apk_path, inner_path)
        return shannon_entropy_bytes(data)
    except Exception:
        return 0.0


def get_file_size_in_apk(apk_path: str, inner_path: str) -> int:
    """Get file size inside APK."""
    try:
        with zipfile.ZipFile(apk_path, "r") as zf:
            info = zf.getinfo(inner_path)
            return info.file_size
    except Exception:
        return 0


# ── Null cleaning ─────────────────────────────────────────────────────────────

def clear_none_recursive(obj):
    """Recursively replace None with proper defaults."""
    if obj is None:
        return ""
    if isinstance(obj, dict):
        return {k: clear_none_recursive(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [clear_none_recursive(item) for item in obj]
    return obj


# ── Resource loading ──────────────────────────────────────────────────────────

def load_permission_mapping(json_path: str) -> dict:
    """Load permission_mapping.json -> {perm_name: level_string}."""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def load_malware_family(excel_path: str) -> dict:
    """Load malradar_reports mapping xlsx -> {report_file: report_family}.
    The xlsx has columns: report_family, report_file, report_in_malradar, ...
    Used as a fallback when report_apk_mappings_new.xlsx has a matching
    report_file but an empty malware_family column.
    Returns {report_file_basename: report_family} dict (key includes extension).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        result = {}
        header = None
        fam_idx = None
        file_idx = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                for i, h in enumerate(header):
                    if "report_family" == h or ("family" in h and "report" in h):
                        fam_idx = i
                    elif "report_file" == h or ("file" in h and "report" in h):
                        file_idx = i
                if file_idx is None or fam_idx is None:
                    wb.close()
                    return {}
                continue
            row_dict = list(row)
            if file_idx < len(row_dict) and fam_idx < len(row_dict):
                file_val = str(row_dict[file_idx]).strip() if row_dict[file_idx] else ""
                fam_val = str(row_dict[fam_idx]).strip() if row_dict[fam_idx] else ""
                if file_val and fam_val:
                    result[file_val] = fam_val
        wb.close()
        return result
    except Exception:
        return {}


def load_sdk_constants(constant_py_path: str) -> tuple:
    """
    Load constant.py's ALL_PACKAGE_SDK and ALL_PACKAGE_TYPE.
    Returns (ALL_PACKAGE_SDK dict, ALL_PACKAGE_TYPE dict).
    """
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location("constant", constant_py_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        sdk = getattr(mod, "ALL_PACKAGE_SDK", {})
        typ = getattr(mod, "ALL_PACKAGE_TYPE", {})
        return sdk, typ
    except Exception:
        return {}, {}


# ── Report behavior → template field mapping ────────────────────────────────

REPORT_BEHAVIOR_TO_FIELD_MAP = [
    ("C2反检测", "c2_encrypted_urls"),
    ("C2 反检测", "c2_encrypted_urls"),
    ("远程控制", "has_c2_communication"),
    ("C2通信", "has_c2_communication"),
    ("C2 通信", "has_c2_communication"),
    ("加密混淆", "encryption_hardcoded_key"),
    ("加密", "encryption_hardcoded_key"),
    ("模拟器对抗", "root_emulator_detection"),
    ("模拟器 对抗", "root_emulator_detection"),
    ("覆盖攻击", "overlay_phishing"),
    ("持久化驻留", "service_keepalive"),
    ("短信劫持", "sms_intercept_via_content_observer"),
    ("短信窃取", "sms_intercept_via_content_observer"),
    ("短信拦截", "sms_intercept_via_broadcast"),
    ("隐私窃取", "device_fingerprint_collection"),
    ("设备信息采集", "device_fingerprint_collection"),
    ("信息窃取", "device_fingerprint_collection"),
    ("网络泄露", "data_exfiltration"),
    ("数据外传", "data_exfiltration"),
    ("数据泄露", "data_exfiltration"),
    ("Root", "root_emulator_detection"),
    ("反检测", "root_emulator_detection"),
    ("保活", "service_keepalive"),
    ("钓鱼", "overlay_phishing"),
    ("overlay", "overlay_phishing"),
    ("设备管理", "admin_abuse_signal"),
    ("系统破坏", "admin_abuse_signal"),
    ("广告欺诈", "ad_click_fraud"),
    ("银行木马", "sms_intercept_via_broadcast"),
    ("蠕虫传播", "sms_delete_capability"),
    ("呼叫转移", "call_forwarding"),
]


def map_report_behavior_to_template_field(report_behavior_name: str) -> str:
    """Map report behavior section name to malicious_behavior template field.
    Longer keys are checked first for more specific matching."""
    normalized = report_behavior_name.replace(" ", "")
    for key, field in REPORT_BEHAVIOR_TO_FIELD_MAP:
        if key in normalized or key in report_behavior_name:
            return field
    return ""


REPORT_BEHAVIOR_TO_KEYWORDS = {
    "信息窃取": ["getDeviceId", "getContentResolver", "Contact", "ContentObserver"],
    "短信劫持": ["SMS_RECEIVED", "abortBroadcast", "content://sms", "ContentObserver"],
    "短信窃取": ["SMS_RECEIVED", "abortBroadcast", "content://sms", "ContentObserver"],
    "短信拦截": ["SMS_RECEIVED", "abortBroadcast"],
    "远程控制": ["HttpURLConnection", "POST", "Socket", "OutputStream"],
    "C2通信": ["HttpURLConnection", "POST", "Socket", "OutputStream"],
    "C2": ["HttpURLConnection", "POST", "Socket", "OutputStream"],
    "广告欺诈": ["dispatchTouchEvent", "performClick", "MotionEvent"],
    "钓鱼": ["TYPE_APPLICATION_OVERLAY", "WindowManager", "SYSTEM_ALERT_WINDOW"],
    "覆盖攻击": ["TYPE_APPLICATION_OVERLAY", "WindowManager", "SYSTEM_ALERT_WINDOW"],
    "overlay": ["TYPE_APPLICATION_OVERLAY", "WindowManager", "SYSTEM_ALERT_WINDOW"],
    "持久化驻留": ["START_STICKY", "AlarmManager", "BOOT_COMPLETED"],
    "保活": ["START_STICKY", "AlarmManager", "BOOT_COMPLETED"],
    "反检测": ["isRooted", "magisk", "google_sdk", "Debug.isDebuggerConnected"],
    "Root": ["isRooted", "magisk", "google_sdk", "su"],
    "模拟器对抗": ["google_sdk", "isEmulator", "Build.FINGERPRINT"],
    "数据外传": ["getBytes", "write", "POST", "putStream", "OutputStream"],
    "网络泄露": ["getBytes", "write", "POST", "putStream", "OutputStream"],
    "数据泄露": ["getBytes", "write", "POST", "putStream", "OutputStream"],
    "加密混淆": ["Cipher.getInstance", "DES", "SecretKeySpec", "decrypt"],
    "加密": ["Cipher.getInstance", "DES", "SecretKeySpec", "decrypt"],
    "蠕虫传播": ["SmsManager", "getAllContacts", "sendTextMessage"],
    "设备管理": ["DeviceAdminReceiver", "BIND_DEVICE_ADMIN", "ACTION_ADD_DEVICE_ADMIN"],
    "系统破坏": ["DeviceAdminReceiver", "BIND_DEVICE_ADMIN", "ACTION_ADD_DEVICE_ADMIN"],
    "银行木马": ["SMS_RECEIVED", "abortBroadcast", "content://sms", "SmsManager"],
    "呼叫转移": ["setCallForward", "CF_ENABLE", "GSM_CALL_FORWARD"],
    "C2反检测": ["Cipher.getInstance", "DES", "Base64", "decrypt"],
}


def map_report_behavior_to_detection_patterns(behavior_name: str) -> list:
    """Return extra detection keyword patterns for a given report behavior name."""
    patterns = []
    for key, pats in REPORT_BEHAVIOR_TO_KEYWORDS.items():
        if key in behavior_name:
            patterns.extend(pats)
    return list(set(patterns))


# ── jadx-gui lifecycle (cross-platform: macOS/Linux/Windows) ──────────────────
#
# On Windows, jadx-gui.exe is an NSIS launcher that forks java.exe/javaw.exe
# then exits immediately, leaving the JVM as an orphan that owns the MCP server.
# On POSIX (macOS/Linux), jadx-gui is a shell script that `exec`s java directly,
# so the Popen handle IS the JVM process (launcher PID == JVM PID).
#
# _jadx_launcher  → Popen handle of jadx-gui launcher
# _jadx_java_pid  → PID of the actual JVM we must kill (may equal launcher PID on POSIX)
# _jadx_apk_name  → basename of opened APK (used to locate process by cmdline)

_IS_WINDOWS = sys.platform == "win32"
_IS_POSIX = sys.platform in ("darwin", "linux")


_jadx_launcher: subprocess.Popen | None = None
_jadx_java_pid: int | None = None
_jadx_apk_name: str = ""


# ── POSIX (macOS / Linux) implementations ─────────────────────────────────────

def _find_jadx_java_child_posix(parent_pid: int, apk_basename: str) -> int | None:
    """POSIX: locate the java process spawned by the jadx-gui launcher.
    On POSIX the jadx-gui shell script uses `exec java`, so the launcher PID
    itself becomes the java process. We also scan for child processes and
    finally fall back to pgrep on the APK basename.
    """
    # 1) Check if the launcher PID is still alive and is a java/jadx process
    try:
        out = subprocess.run(
            ["ps", "-p", str(parent_pid), "-o", "comm="],
            capture_output=True, text=True, timeout=5,
        ).stdout.strip()
        if out and ("java" in out.lower() or "jadx" in out.lower()):
            return parent_pid
    except Exception:
        pass

    # 2) Look for child processes of the launcher (ppid == parent_pid)
    try:
        out = subprocess.run(
            ["ps", "-o", "pid,ppid,command"],
            capture_output=True, text=True, timeout=5,
        ).stdout
        for line in out.strip().splitlines()[1:]:  # skip header
            parts = line.strip().split(None, 2)
            if len(parts) >= 3:
                try:
                    pid = int(parts[0])
                    ppid = int(parts[1])
                except ValueError:
                    continue
                cmd = parts[2]
                if ppid == parent_pid and ("java" in cmd.lower() or "jadx" in cmd.lower()):
                    return pid
    except Exception:
        pass

    # 3) Fallback: find java process whose cmdline contains the APK basename
    try:
        out = subprocess.run(
            ["pgrep", "-f", apk_basename],
            capture_output=True, text=True, timeout=10,
        ).stdout
        for word in out.split():
            try:
                return int(word.strip())
            except ValueError:
                pass
    except Exception:
        pass

    return None


def _kill_all_jadx_orphans_posix():
    """POSIX: kill all existing jadx-gui / java processes running jadx."""
    try:
        out = subprocess.run(
            ["pgrep", "-f", "jadx"],
            capture_output=True, text=True, timeout=10,
        ).stdout
        for word in out.split():
            try:
                pid = int(word.strip())
                _kill_java_pid(pid)
                print(f"[+] Pre-launch killed orphan jadx JVM PID={pid}")
            except (ValueError, Exception):
                pass
    except Exception:
        pass


def _kill_java_pid_posix(pid: int) -> bool:
    """POSIX: terminate then force-kill a single process via SIGTERM → SIGKILL."""
    try:
        os.kill(pid, signal.SIGTERM)
    except ProcessLookupError:
        return True  # already gone
    except Exception:
        pass

    # Wait briefly for graceful shutdown
    time.sleep(1)

    # Check if still alive
    try:
        out = subprocess.run(
            ["ps", "-p", str(pid), "-o", "pid="],
            capture_output=True, text=True, timeout=3,
        ).stdout.strip()
        if not out:
            return True  # gone
    except Exception:
        pass

    # Force kill
    try:
        os.kill(pid, signal.SIGKILL)
        return True
    except ProcessLookupError:
        return True
    except Exception:
        return False


def _kill_remaining_java_by_cmdline_posix() -> int:
    """POSIX: kill any java process whose cmdline contains our APK basename."""
    killed = 0
    try:
        out = subprocess.run(
            ["pgrep", "-f", _jadx_apk_name],
            capture_output=True, text=True, timeout=10,
        ).stdout
        for word in out.split():
            try:
                pid = int(word.strip())
                if _kill_java_pid(pid):
                    killed += 1
                    print(f"[+] Killed remaining java process PID={pid}")
            except (ValueError, Exception):
                pass
    except Exception:
        pass
    return killed


# ── Windows implementations (kept for cross-platform support) ────────────────

def _find_jadx_java_child_win(parent_pid: int, apk_basename: str) -> int | None:
    """Windows: locate the java.exe/javaw.exe child process via tasklist/powershell."""
    import re as _re
    # Stage 1: query children of the launcher via tasklist parent-pid filter
    try:
        out = subprocess.run(
            ["tasklist", "/FI", f"PID eq {parent_pid}", "/FO", "CSV", "/NH"],
            capture_output=True, text=True, timeout=5,
        ).stdout
        for line in out.strip().splitlines():
            parts = line.strip().strip('"').split('","')
            if len(parts) >= 2:
                try:
                    return int(parts[0])
                except ValueError:
                    pass
    except Exception:
        pass

    # Stage 2: enumerate all java processes matching APK in cmdline
    apk_esc = _re.escape(apk_basename)
    try:
        out = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f'$procs = Get-CimInstance Win32_Process -Filter "name=\'javaw.exe\'" | '
             f'Where-Object {{ $_.CommandLine -match \'{apk_esc}\' }}; '
             f'if ($procs) {{ $procs.ProcessId }} else {{ \'\'}}, '
             f'$procs2 = Get-CimInstance Win32_Process -Filter "name=\'java.exe\'" | '
             f'Where-Object {{ $_.CommandLine -match \'{apk_esc}\' }}; '
             f'if ($procs2) {{ $procs2.ProcessId }}'],
            capture_output=True, text=True, timeout=10,
        ).stdout
        for word in out.split():
            try:
                return int(word.strip('"'))
            except ValueError:
                pass
    except Exception:
        pass
    return None


def _kill_all_jadx_orphans_win():
    """Windows: kill all existing java.exe/javaw.exe running jadx-gui."""
    try:
        out = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             '$procs = Get-CimInstance Win32_Process -Filter "name=\'javaw.exe\'" | '
             'Where-Object { $_.CommandLine -match "jadx" } | '
             'Select-Object -ExpandProperty ProcessId; '
             'if ($procs) { $procs } else { "" }'],
            capture_output=True, text=True, timeout=10,
        ).stdout
        for word in out.split():
            try:
                pid = int(word.strip('"'))
                _kill_java_pid(pid)
                print(f"[+] Pre-launch killed orphan jadx JVM PID={pid}")
            except (ValueError, Exception):
                pass
    except Exception:
        pass


def _kill_java_pid_win(pid: int) -> bool:
    """Windows: terminate then force-kill via taskkill."""
    try:
        proc = subprocess.Popen(
            ["taskkill", "/PID", str(pid)],
            creationflags=subprocess.CREATE_NO_WINDOW,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        proc.wait(timeout=5)
        check = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/FO", "CSV", "/NH"],
            capture_output=True, text=True, timeout=5,
        )
        if not check.stdout.strip():
            return True
    except Exception:
        pass
    try:
        subprocess.run(
            ["taskkill", "/F", "/PID", str(pid)],
            capture_output=True, timeout=8,
        )
        return True
    except Exception:
        return False


def _kill_remaining_java_by_cmdline_win() -> int:
    """Windows: kill leftover jadx JVMs.

    The tracked-PID kill misses JVMs from PREVIOUS APKs of this agent (the
    .bat launcher detaches javaw; the same-APC-name matcher only sees the
    CURRENT APK). Broader rule: enumerate ALL java/javaw running the jadx
    jar and kill every one that is NOT a live jadx-gui instance.

    "Live" is determined precisely: each agent's jadx-gui serves its MCP
    plugin HTTP port (8650 + AGENT_ID). We resolve the JVM PIDs bound to
    any of those ports via netstat — those are in active use by some agent
    (ours or a sibling's) and are exempt; everything else running the jadx
    jar is garbage from a previous APK and gets killed.
    """
    import re as _re
    import json as _json
    live_pids = _live_jadx_plugin_pids()
    killed = 0
    for image in ("java.exe", "javaw.exe"):
        try:
            ps_cmd = (
                "Get-CimInstance Win32_Process -Filter \"name='" + image + "'\" | "
                "Where-Object { $_.CommandLine -match 'jadx-.*-all\\.jar' } | "
                "Select-Object ProcessId, CommandLine | "
                "ConvertTo-Json -Compress"
            )
            out = subprocess.run(
                ["powershell", "-NoProfile", "-Command", ps_cmd],
                capture_output=True, text=True, timeout=15,
            ).stdout
            try:
                entries = _json.loads(out)
            except Exception:
                entries = []
            if isinstance(entries, dict):
                entries = [entries]
            for e in entries:
                try:
                    pid = int(e["ProcessId"])
                    cmdline = e.get("CommandLine") or ""
                    if pid in live_pids:
                        continue  # actively serving an agent's MCP port
                    m = _re.search(r"([0-9a-fA-F]{64})\.apk", cmdline)
                    tag = m.group(1)[:12] if m else "?"
                    if _kill_java_pid(pid):
                        killed += 1
                        print(f"[+] Killed leftover jadx JVM PID={pid} (APK {tag})")
                except (ValueError, KeyError, Exception):
                    pass
        except Exception:
            pass
    return killed


def _live_jadx_plugin_pids() -> set:
    """PIDs of JVMs listening on any jadx-ai-mcp plugin port (8650..8699).

    Each pipeline agent's jadx-gui binds its own port (config:
    JADX_GUI_PLUGIN_PORT = 8650 + AGENT_ID). A leftover jadx-gui from a
    previous APK is either dead or unbound (its port was taken over by the
    next launch), so "bound to a plugin port" reliably identifies JVMs in
    active use by SOME agent — safe to exempt from the sweep.
    """
    pids = set()
    try:
        # netstat output is CP936/GBK on zh-CN Windows — text=True with the
        # default UTF-8 codec raises UnicodeDecodeError and returns None.
        # Decode bytes ourselves, tolerantly.
        raw = subprocess.run(
            ["netstat", "-ano", "-p", "TCP"],
            capture_output=True, timeout=10,
        ).stdout or b""
        out = raw.decode("utf-8", errors="replace")
        for line in out.splitlines():
            if "LISTENING" not in line.upper():
                continue
            m = re.search(r":(86[5-9]\d)\s+\S+\s+LISTENING\s+(\d+)", line, re.IGNORECASE)
            if m:
                pids.add(int(m.group(2)))
    except Exception:
        pass
    return pids


# ── Cross-platform dispatchers ────────────────────────────────────────────────

def _find_jadx_java_child(parent_pid: int, apk_basename: str) -> int | None:
    """Locate the JVM process spawned by the jadx-gui launcher (cross-platform)."""
    if _IS_WINDOWS:
        return _find_jadx_java_child_win(parent_pid, apk_basename)
    return _find_jadx_java_child_posix(parent_pid, apk_basename)


def _kill_all_jadx_orphans():
    """Kill ALL existing jadx-gui JVM processes (cross-platform)."""
    if _IS_WINDOWS:
        _kill_all_jadx_orphans_win()
    else:
        _kill_all_jadx_orphans_posix()


def _kill_java_pid(pid: int) -> bool:
    """Terminate then force-kill a single process (cross-platform)."""
    if _IS_WINDOWS:
        return _kill_java_pid_win(pid)
    return _kill_java_pid_posix(pid)


def _kill_remaining_java_by_cmdline() -> int:
    """Brute-force: kill any java process whose cmdline contains our APK (cross-platform)."""
    if _IS_WINDOWS:
        return _kill_remaining_java_by_cmdline_win()
    return _kill_remaining_java_by_cmdline_posix()


def _set_jadx_plugin_port(port: int):
    """Write the jadx AI MCP plugin port via java.util.prefs (a fresh JVM reads it
    reliably; `defaults write` targets a different plist path and does not reach
    Java's Preferences API on macOS)."""
    try:
        from config import JAVA_BIN, JADX_PREFS_JAR
        subprocess.run(
            [JAVA_BIN, "-jar", JADX_PREFS_JAR, str(port)],
            capture_output=True, text=True, timeout=15,
        )
    except Exception:
        pass


def _wait_plugin_health(port: int, timeout: int = 180, interval: int = 2) -> bool:
    """Poll the jadx AI MCP plugin health endpoint until ready or timeout."""
    try:
        import requests as _req
    except ImportError:
        return False
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            r = _req.get(f"http://127.0.0.1:{port}/health", timeout=2)
            if r.status_code == 200:
                print(f"  [+] jadx plugin health check passed on port {port}")
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


def _acquire_file_lock(lock_path: str):
    """Acquire an exclusive file lock. Returns the open file object (or None if unavailable).
    POSIX: fcntl.flock. Windows: msvcrt.locking."""
    try:
        f = open(lock_path, "a+")
        if sys.platform == "win32":
            import msvcrt
            f.seek(0)
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(f.fileno(), fcntl.LOCK_EX)
        return f
    except Exception:
        try:
            f.close()
        except Exception:
            pass
        return None


def _release_file_lock(lock):
    """Release a file lock previously acquired by _acquire_file_lock."""
    if lock is None:
        return
    try:
        if sys.platform == "win32":
            import msvcrt
            lock.seek(0)
            msvcrt.locking(lock.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl
            fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
    except Exception:
        pass
    try:
        lock.close()
    except Exception:
        pass


def launch_jadx_gui(jadx_path: str, apk_path: str, port: int = None) -> subprocess.Popen:
    """Launch jadx-gui with the target APK; capture the JVM child PID.
    Parallel-safe: serializes launches across agents via a file lock so the Java
    Preferences port can be set race-free before jadx-gui reads it. Does NOT kill
    other agents' jadx-gui instances."""
    global _jadx_launcher, _jadx_java_pid, _jadx_apk_name

    if not os.path.exists(jadx_path):
        raise FileNotFoundError(f"jadx-gui not found at: {jadx_path}")

    from config import JADX_GUI_PLUGIN_PORT, JADX_LAUNCH_LOCK_FILE
    if port is None:
        port = JADX_GUI_PLUGIN_PORT

    abs_apk = os.path.abspath(apk_path)
    _jadx_apk_name = os.path.basename(abs_apk)
    _jadx_java_pid = None

    # Serialize launches: set plugin port then launch while holding the lock so no
    # other agent overwrites the Preferences port before this instance binds.
    lock = _acquire_file_lock(JADX_LAUNCH_LOCK_FILE)
    try:
        _set_jadx_plugin_port(port)

        cmd = [jadx_path, abs_apk]
        if _IS_WINDOWS:
            _jadx_launcher = subprocess.Popen(cmd, creationflags=subprocess.CREATE_NO_WINDOW)
        else:
            # POSIX: start_new_session detaches the child into its own process group,
            # so we can later kill the entire group without affecting this Python process.
            _jadx_launcher = subprocess.Popen(cmd, start_new_session=True)
        launcher_pid = _jadx_launcher.pid
        from config import JADX_LAUNCH_TIMEOUT
        print(f"[+] jadx-gui launched (launcher PID={launcher_pid}, port={port}): {abs_apk}")

        # Wait for the plugin to bind on our port while holding the lock.
        _wait_plugin_health(port, timeout=JADX_LAUNCH_TIMEOUT)

        time.sleep(2)
        _jadx_java_pid = _find_jadx_java_child(launcher_pid, _jadx_apk_name)
        if _jadx_java_pid:
            print(f"[+] Captured jadx JVM child PID={_jadx_java_pid}")
        else:
            print("[!] Could not capture jadx JVM PID — will use process-group kill on cleanup")
    finally:
        _release_file_lock(lock)

    return _jadx_launcher


def kill_jadx_gui():
    """Terminate the jadx-gui JVM properly (cross-platform).
    Parallel-safe: only kills this process's own launcher / tracked JVM PID, never
    touches other agents' jadx-gui instances (no global pgrep sweeps)."""
    global _jadx_launcher, _jadx_java_pid, _jadx_apk_name

    if _jadx_apk_name:  # something was started
        # 1) Terminate the tracked JVM child first (most reliable)
        if _jadx_java_pid:
            print(f"[*] Killing tracked jadx JVM PID={_jadx_java_pid}")
            _kill_java_pid(_jadx_java_pid)

        # 2) Kill the launcher's own process group (launcher + any children)
        if _jadx_launcher:
            try:
                pgid = os.getpgid(_jadx_launcher.pid)
                if pgid and pgid != os.getpgrp():
                    print(f"[*] Killing jadx process group PID={pgid}")
                    os.killpg(pgid, signal.SIGTERM)
                    time.sleep(1)
            except Exception:
                pass
            if _jadx_launcher.poll() is None:
                try:
                    _jadx_launcher.terminate()
                    _jadx_launcher.wait(timeout=3)
                except Exception:
                    try:
                        _jadx_launcher.kill()
                    except Exception:
                        pass

        if not _jadx_java_pid:
            print("[!] No jadx JVM process found — may already be dead")

        # 3) FINAL SAFETY NET (Windows-critical): kill any java/javaw whose
        # command line still references THIS agent's APK. On Windows the
        # .bat launcher uses `start /B javaw`, detaching the JVM from the
        # launcher — the captured "child PID" can point at the wrong process,
        # leaving orphan JVMs (~300MB each). Matching on our own APK name
        # keeps parallel agents safe (each agent's APK differs).
        try:
            killed = _kill_remaining_java_by_cmdline()
            if killed:
                print(f"[+] cmdline sweep killed {killed} leftover jadx JVM(s)")
        except Exception as e:
            print(f"[!] cmdline sweep failed (non-fatal): {e}")

        print("[+] jadx-gui cleanup complete")

    # Reset
    _jadx_launcher = None
    _jadx_java_pid = None
    _jadx_apk_name = ""


def clear_jadx_cache():
    """Clear jadx plugin cache for next APK via HTTP. Tolerate dead plugin gracefully."""
    try:
        from config import JADX_GUI_PLUGIN_PORT
        import requests as _req
        r = _req.post(f"http://127.0.0.1:{JADX_GUI_PLUGIN_PORT}/clear-cache", timeout=5)
        if r.status_code == 200:
            print("[+] JADX cache cleared")
        else:
            print(f"[!] JADX cache clear returned {r.status_code}")
    except ImportError:
        print("[!] JADX cache clear skipped (requests not available)")
    except Exception as e:
        # plugin may already be dead after kill_jadx_gui — that's expected
        print(f"[!] JADX cache clear skipped (plugin may be dead): {e}")


# ── JSON helpers ──────────────────────────────────────────────────────────────

def save_json(obj, file_path: str):
    """Save object to JSON file with UTF-8 encoding."""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def load_json(file_path: str) -> object:
    """Load JSON file. Returns empty dict on error."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


# ── Report mapping (Excel → MD report → known behaviors) ───────────────────────

def load_apk_report_mapping(sha256: str, excel_path: str, report_md_dir: str) -> tuple:
    """
    Query report_apk_mappings_new.xlsx by sha256.
    Returns (report_md_path, malware_family) or (None, "").
    report_md_path: full path to the .md file in report_md_dir, or None if not found.
    malware_family: string from excel column "malware_family", or "" if not found.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        header = None
        report_file = None
        malware_family = ""
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                col_map = {}
                for i, h in enumerate(header):
                    if h == "report_file":
                        col_map["report_file"] = i
                    elif h == "sha256":
                        col_map["sha256"] = i
                    elif h == "malware_family":
                        col_map["malware_family"] = i
                if "sha256" not in col_map:
                    wb.close()
                    return None, ""
                continue
            row_sha = str(row[col_map["sha256"]]).strip() if col_map["sha256"] < len(row) and row[col_map["sha256"]] else ""
            if row_sha == sha256:
                if "report_file" in col_map and col_map["report_file"] < len(row):
                    report_file = str(row[col_map["report_file"]]).strip()
                else:
                    report_file = None
                if "malware_family" in col_map and col_map["malware_family"] < len(row):
                    malware_family = str(row[col_map["malware_family"]]).strip()
                else:
                    malware_family = ""
                break
        wb.close()

        # Build full MD path
        md_path = None
        if report_file and os.path.isdir(report_md_dir):
            md_name = os.path.splitext(report_file)[0] + ".md"
            md_path = os.path.join(report_md_dir, md_name)
            if not os.path.isfile(md_path):
                md_path = None

        return md_path, malware_family
    except Exception:
        return None, ""


def parse_report_behaviors(report_md_path: str) -> tuple:
    """
    Parse report MD file's '## 3. 恶意行为描述' section.
    Returns (behavior_list, behavior_keywords, behavior_techs).
    behavior_list: [{"name": ..., "risk": ...}, ...]
    behavior_keywords: ["keyword1", "keyword2", ...]
    behavior_techs: [{"name": ..., "risk": ..., "techniques": [...]}, ...]
    """
    behaviors = []
    keywords = []
    behavior_techs = []
    try:
        with open(report_md_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Find "## 3. 恶意行为描述" section
        section_idx = content.find("## 3. 恶意行为描述")
        if section_idx == -1:
            section_idx = content.find("## 3.")
        if section_idx == -1:
            return [], [], []

        section = content[section_idx:]
        # Find end of section (next ## or EOF)
        next_section = section.find("\n## ", 3)
        if next_section != -1:
            section = section[:next_section]

        import re
        # Extract ### 3.X 行为名 (**等级**) and corresponding techniques
        behavior_pattern = re.compile(r"###\s+3\.\d+\s+(.+?)\s+\(\*\*?(\w+)\*\*?\)")
        tech_inline_pattern = re.compile(r"\*\*涉及技术\*\*:\s*(.*?)(?:\n|$)")

        for m in behavior_pattern.finditer(section):
            behavior_name = m.group(1).strip()
            risk_level = m.group(2).strip()
            behaviors.append({"name": behavior_name, "risk": risk_level})

            # Find techniques in the block following this behavior header
            start = m.end()
            next_behavior = list(behavior_pattern.finditer(section))
            end = len(section)
            for nb in next_behavior:
                if nb.start() > start:
                    end = nb.start()
                    break
            block = section[start:end]

            tech_matches = tech_inline_pattern.findall(block)
            techs = []
            for tm in tech_matches:
                for t in tm.strip().split(","):
                    t = t.strip()
                    if t:
                        techs.append(t)

            behavior_techs.append({
                "name": behavior_name,
                "risk": risk_level,
                "techniques": techs,
            })

        # Extract keywords from "涉及技术" lines globally
        tech_pattern = re.compile(r"\*\*涉及技术\*\*:\s*(.+)")
        for m in tech_pattern.finditer(section):
            techs = m.group(1).strip().split(",")
            for t in techs:
                t = t.strip()
                if t:
                    keywords.append(t)

        # Extract specific command names (sms_send, sms_grab, etc.)
        cmd_pattern = re.compile(r"\b(\w+_\w+)\b")
        cmd_matches = cmd_pattern.findall(section)
        for cm in cmd_matches:
            if cm not in keywords and len(cm) > 4:
                keywords.append(cm)

    except Exception:
        pass

    return behaviors, keywords, behavior_techs
